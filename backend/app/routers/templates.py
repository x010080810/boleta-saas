from fastapi import APIRouter
from fastapi.responses import FileResponse
import os
import tempfile
import openpyxl
from app.core.config import settings

router = APIRouter()


def generate_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EMPLEADOS"

    headers = [
        "tipo_documento",
        "numero_documento",
        "apellidos_nombres",
        "email",
        "cargo",
        "fecha_ingreso",
        "dias_laborados",
        "asignacion_familiar",
        "ING_haber_basico",
        "ING_asignacion_familiar_monto",
        "ING_vacaciones",
        "ING_horas_extras",
        "ING_bono_productividad",
        "ING_gratificacion",
        "ING_cts",
        "ING_feriado",
        "ING_bono_trimestral",
        "ING_otros",
        "DESC_afp_obligatorio",
        "DESC_comision_afp",
        "DESC_prima_seguro",
        "DESC_onp",
        "DESC_renta_5ta",
        "DESC_otros",
        "APOR_essalud",
        "APOR_credito_eps",
        "APOR_vida_ley",
        "APOR_otros",
        "total_ingresos",
        "total_descuentos",
        "neto_pagar",
        "neto_pagar_usd",
    ]

    ws.append(headers)

    ws.append([
        "01", "12345678", "EJEMPLO APELLIDO NOMBRE", "correo@ejemplo.com",
        "CARGO EJEMPLO", "2024-01-15", 30, "NO",
        2500, 113, 0, 0, 200, 0, 0, 0, 0, 0,
        250, 0, 20, 0, 0, 0,
        225, 0, 5, 0,
        2813, 270, 2543, 0,
    ])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 20

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)
    return temp.name


@router.get("/excel")
async def download_excel_template():
    filepath = generate_excel_template()
    return FileResponse(
        path=filepath,
        filename="plantilla_boletas.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
